"""
CS506 : parceExcel
Team : Vidya Akavoor, Lauren DiSalvo, Sreeja Keesara
Description :

Notes : 

October 28, 2018
"""

import xlrd
import xlwt
from Classes import Bus, School, Student
from openpyxl import Workbook

TEST_BUS_DICT = {}
TEST_SCHOOL_DICT = {}
TEST_STUDENT_DICT = {}

# workbook = xlrd.open_workbook('2016-2017 Framingham Bus Data.xlsx')
# worksheet = workbook.sheet_by_name('qmf_temp')
#
# studentResidentialAddresses = worksheet.col_values(2)
# bus = Bus(bus_id=5, capacity=60)
# print(bus.latitude)


def create_template(busDict, studentDict, schoolDict):
    workbook = Workbook()

    bus_sheet = workbook.create_sheet('Buses')
    bus_sheet.append(('Bus Id', 'Bus Capacity', 'Bus Latitude', 'Bus Lognitude', 'Bus Type', 'Bus Yard', 'Bus Yard Address'))

    stop_sheet = workbook.create_sheet('Stop-Assignments')
    stop_sheet.append(('Student Id', 'Student Latitude', 'Student Lognitude', 'School Latitude', 'School Longitude', 'Bus ID', 'Stop Latitude', 'Stop Longitude'))

    routes_sheet = workbook.create_sheet('Routes')
    routes_sheet.append(('Bus ID', 'Waypoint Latitude', 'Waypoint Longitude', 'Waypoint Address'))

    workbook.save('current_bus_routes.xlsx')


create_template(TEST_BUS_DICT, TEST_STUDENT_DICT, TEST_SCHOOL_DICT)





